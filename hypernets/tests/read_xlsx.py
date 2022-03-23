
import os
import xlrd


basedir = "/mnt/c/Users/wuhf/Desktop/hyperctl/demo/batches-data/batch-openml-cpu"

for file_ in os.listdir(basedir):
    abs_file_ = os.path.join(basedir, file_, 'report.xlsx')
    # print(abs_file_)
    if os.path.exists(abs_file_):
        print(abs_file_)



#%%

filepath = "/mnt/c/Users/wuhf/Desktop/hyperctl/demo/batches-data/batch-openml-cpu/7hhDFNsw/report.xlsx"
import pandas as pd
df = pd.read_excel(filepath, sheet_name="Evaluation")
print(df)


#%%

from openpyxl import load_workbook

wb = load_workbook('data1.xlsx')
sheets = wb.worksheets  # 获取当前所有的sheet
print(sheets)

# 获取第一张sheet
sheet1 = sheets[0]
# sheet1 = wb['Sheet']  # 也可以通过已知表名获取sheet
print(sheet1)

# 通过Cell对象读取
cell_11 = sheet1.cell(1, 1).value
print(cell_11)
cell_11 = sheet1.cell(1, 2).value
print(cell_11)